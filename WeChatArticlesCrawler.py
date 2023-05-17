"""
网页爬取「fakied」、「token」、「cookie」、「user-agent」
Fiddler爬取「Key」、「Pass_ticket」、「appmasg_token」、「Cookie」、「User-Agent」

Step1:利用微信公众号平台获取指定公众号的文章列表
Step2：模拟微信访问文章并提取信息
"""

import requests
import time
import json
from openpyxl import Workbook
import random
import re

url = "https://mp.weixin.qq.com/cgi-bin/appmsg"
Cookie = "从微信公众平台获取到的cookie" # 从微信公众平台获取的Cookie 标头
headers = {
    "Cookie": Cookie,
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36 Edg/113.0.1774.35",
}
token = "XXX" # 从微信公众平台获取的token 负载
fakeid = "XXX" # 从微信公众平台获取的fakeid 负载
type = '9'
data1 = {
    "token": token,
    "lang": "zh_CN",
    "f": "json",
    "ajax": "1",
    "action": "list_ex",
    "begin": "0",
    "count": "4",
    "query": "",
    "fakeid": fakeid,
    "type": type,
}

def getMoreInfo(link):
    # 获得mid,_biz,idx,sn 这几个在link中的信息
    mid = link.split("&")[1].split("=")[1]
    idx = link.split("&")[2].split("=")[1]
    sn = link.split("&")[3].split("=")[1]
    _biz = link.split("&")[0].split("_biz=")[1]
    pass_ticket = "fiddler获取" # 从fiddler中获取
    appmsg_token = "fiddler获取" # 从fiddler中获取
    url = "http://mp.weixin.qq.com/mp/getappmsgext" # 获取详情页的网址
    phoneCookie = "fiddler获取" # 从fiddler获取
    headers = {
        "Cookie": phoneCookie,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36 NetType/WIFI MicroMessenger/7.0.20.1781(0x6700143B) WindowsWechat(0x6309001c) XWEB/6763"
    }
    data = {
        "is_only_read": "1",
        "is_temp_url": "0",
        "appmsg_type": "9",
        'reward_uin_count': '0'
    }
    params = {
        "__biz": _biz,
        "mid": mid,
        "sn": sn,
        "idx": idx,
        "key": "fiddler获取", # 从fiddler中获取
        "pass_ticket": pass_ticket,
        "appmsg_token": appmsg_token,
        "uin": "MjgwMjI0MTMyNQ==",
        "wxtoken": "777",
    }

    requests.packages.urllib3.disable_warnings()

    # 使用post方法进行提交，这里返回了一个json，里面是单个文章的数据
    content = requests.post(url, headers=headers, data=data, params=params).json()

    # 在上面返回的json中获取并输出阅读数喜爱数在读数
    try:
        readNum = content["appmsgstat"]["read_num"]
        print("阅读数:"+str(readNum))
    except:
        readNum = 0
    try:
        likeNum = content["appmsgstat"]["like_num"]
        print("喜爱数:"+str(likeNum))
    except:
        likeNum = 0
    try:
        old_like_num = content["appmsgstat"]["old_like_num"]
        print("在读数:"+str(old_like_num))
        print(" ")
    except:
        old_like_num = 0
    time.sleep(3) # 歇3s，防止被封
    return readNum, likeNum,old_like_num,footing_result
def getAllInfo(url):
    # 拿一页，存一页
    messageAllInfo = []
    # begin 从0开始
    for i in range(33): # 设置爬虫页码
        begin = i * 4
        data1["begin"] = begin
        requests.packages.urllib3.disable_warnings()
        content_json = requests.get(url, headers=headers, params=data1, verify=False).json()
        time.sleep(random.randint(1, 10))
        if "app_msg_list" in content_json:
            for item in content_json["app_msg_list"]:
                timestamp = item['create_time']
                time_local = time.localtime(timestamp)
                spider_url = item['link']
                readNum, likeNum,old_like_num = getMoreInfo(spider_url)
                info = {
                    "title": item['title'],
                    "createTime": time.strftime("%Y-%m-%d %H:%M:%S",time_local),
                    "url": item['link'],
                    "readNum": readNum,
                    "likeNum": likeNum,
                    "old_like_num": str(old_like_num),
                }
                messageAllInfo.append(info)
    return messageAllInfo
def main():
    f = Workbook()  # 创建一个workbook 设置编码
    sheet = f.active # 创建sheet表单
    # 写入表头
    sheet.cell(row=1, column=1).value = 'title(推文标题)'  # 第一行第一列
    sheet.cell(row=1, column=2).value = 'creatTime(发布时间)'
    sheet.cell(row=1, column=3).value = 'url(推文链接)'
    sheet.cell(row=1, column=4).value = 'readNum(阅读数)'
    sheet.cell(row=1, column=5).value = 'likeNum(喜爱数)'
    sheet.cell(row=1, column=6).value = 'old_like_num(在看数)'
    sheet.cell(row=1, column=7).value = 'footing_result(小尾巴)'
    messageAllInfo = getAllInfo(url) # 获取信息
    print(messageAllInfo)
    print(len(messageAllInfo)) # 输出列表长度
    # 写内容
    for i in range(1, len(messageAllInfo)+1):
        sheet.cell(row=i + 1, column=1).value = messageAllInfo[i - 1]['title']
        sheet.cell(row=i + 1, column=2).value = messageAllInfo[i - 1]['createTime']
        sheet.cell(row=i + 1, column=3).value = messageAllInfo[i - 1]['url']
        sheet.cell(row=i + 1, column=4).value = messageAllInfo[i - 1]['readNum']
        sheet.cell(row=i + 1, column=5).value = messageAllInfo[i - 1]['likeNum']
        sheet.cell(row=i + 1, column=6).value = messageAllInfo[i - 1]['old_like_num']
    f.save(u'result.xls')  # 保存文件
if __name__ == '__main__':
    main()
