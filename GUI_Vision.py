import requests
from bs4 import BeautifulSoup
import time
import random
from openpyxl import Workbook
import tkinter as tk
from tkinter import scrolledtext
import sys


def getMoreInfo(link):
    # 获得mid,_biz,idx,sn 这几个在link中的信息
    mid = link.split("&")[1].split("=")[1]
    idx = link.split("&")[2].split("=")[1]
    sn = link.split("&")[3].split("=")[1]
    _biz = link.split("&")[0].split("_biz=")[1]
    pass_ticket = WX_Pass_ticket
    appmsg_token = WX_Appmsg_Token
    url = "http://mp.weixin.qq.com/mp/getappmsgext" # 获取详情页的网址
    phoneCookie = WX_Cookie
    headers = {
        "Cookie": phoneCookie,
        "User-Agent": WX_User_Agent
    }
    data = {
        "is_only_read": "1",
        "is_temp_url": "0",
        "appmsg_type": "9",
        'reward_uin_count': '0'
    }
    params = {
        "__biz": _biz,
        "mid": mid,
        "sn": sn,
        "idx": idx,
        "key": WX_Key,
        "pass_ticket": pass_ticket,
        "appmsg_token": appmsg_token,
        "uin": "MjgwMjI0MTMyNQ==",
        "wxtoken": "777",
    }


    requests.packages.urllib3.disable_warnings()
    # 使用post方法进行提交，这里返回了一个json，里面是单个文章的数据
    content = requests.post(url, headers=headers, data=data, params=params).json()

    # 在上面返回的json中获取并输出阅读数喜爱数在读数
    try:
        readNum = content["appmsgstat"]["read_num"]
        print("阅读数:" + str(readNum))
    except:
        readNum = 0
    try:
        likeNum = content["appmsgstat"]["like_num"]
        print("喜爱数:" + str(likeNum))
    except:
        likeNum = 0
    try:
        old_like_num = content["appmsgstat"]["old_like_num"]
        print("在读数:" + str(old_like_num))
    except:
        old_like_num = 0

    time.sleep(3) # 歇3s，防止被封
    return readNum, likeNum, old_like_num

def getAllInfo():
    url = "https://mp.weixin.qq.com/cgi-bin/appmsg"
    Cookie = platform_Cookie
    headers = {
        "Cookie": Cookie,
        "User-Agent": platform_UserAgent
    }
    token = platform_Token  # 公众号
    fakeid = platform_Fakeid  # 公众号对应的id
    type = '9'
    data1 = {
        "token": token,
        "lang": "zh_CN",
        "f": "json",
        "ajax": "1",
        "action": "list_ex",
        "begin": "0",
        "count": "4",
        "query": "",
        "fakeid": fakeid,
        "type": type,
    }

    # 拿一页，存一页
    messageAllInfo = []
    # begin 从0开始
    for i in range(66):  # 设置爬虫页码
        begin = i * 4
        data1["begin"] = begin
        requests.packages.urllib3.disable_warnings()
        content_json = requests.get(url, headers=headers, params=data1, verify=False).json()
        time.sleep(random.randint(1, 10))
        if "app_msg_list" in content_json:
            for item in content_json["app_msg_list"]:
                timestamp = item['create_time']
                time_local = time.localtime(timestamp)
                spider_url = item['link']
                readNum, likeNum,old_like_num = getMoreInfo(spider_url)
                info = {
                    "title": item['title'],
                    "createTime": time.strftime("%Y-%m-%d %H:%M:%S",time_local),
                    "url": item['link'],
                    "readNum": readNum,
                    "likeNum": likeNum,
                    "old_like_num": old_like_num,
                }
                messageAllInfo.append(info)
    return messageAllInfo

def main():
    f = Workbook()  # 创建一个workbook 设置编码
    sheet = f.active  # 创建sheet表单
    # 写入表头
    sheet.cell(row=1, column=1).value = 'title(推文标题)'  # 第一行第一列
    sheet.cell(row=1, column=2).value = 'creatTime(发布时间)'
    sheet.cell(row=1, column=3).value = 'url(推文链接)'
    sheet.cell(row=1, column=4).value = 'readNum(阅读数)'
    sheet.cell(row=1, column=5).value = 'likeNum(喜爱数)'
    sheet.cell(row=1, column=6).value = 'old_like_num(在看数)'
    messageAllInfo = getAllInfo()   # 获取信息
    # print(messageAllInfo)
    print(len(messageAllInfo))  # 输出列表长度
    # 写内容
    for i in range(1, len(messageAllInfo)+1):
        sheet.cell(row=i + 1, column=1).value = messageAllInfo[i - 1]['title']
        sheet.cell(row=i + 1, column=2).value = messageAllInfo[i - 1]['createTime']
        sheet.cell(row=i + 1, column=3).value = messageAllInfo[i - 1]['url']
        sheet.cell(row=i + 1, column=4).value = messageAllInfo[i - 1]['readNum']
        sheet.cell(row=i + 1, column=5).value = messageAllInfo[i - 1]['likeNum']
        sheet.cell(row=i + 1, column=6).value = messageAllInfo[i - 1]['old_like_num']
    f.save(u'爬取结果.xls')  # 保存文件


class GUI:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Wx文章信息抓取器 v23.6.10 ByLydia")

        self.labels = ["platform_UserAgent", "platform_Cookie", "platform_Token", "platform_Fakeid",
                       "WX_User_Agent", "WX_Cookie", "WX_Appmsg_Token", "WX_Pass_ticket", "WX_Key"]
        self.entries = {}

        for idx, label in enumerate(self.labels, start=1):
            tk.Label(self.window, text=label).grid(row=idx, column=0)
            self.entries[label] = tk.Entry(self.window)
            self.entries[label].grid(row=idx, column=1)

        self.btn_submit = tk.Button(self.window, text="提交并运行", command=self.run_main)
        self.btn_submit.grid(row=len(self.labels) + 1, column=0, columnspan=2)

        self.output_area = scrolledtext.ScrolledText(self.window)
        self.output_area.grid(row=len(self.labels) + 2, column=0, columnspan=2)

        self.window.mainloop()

    def run_main(self):
        global platform_UserAgent, platform_Cookie, platform_Token, platform_Fakeid
        global WX_User_Agent, WX_Cookie, WX_Appmsg_Token, WX_Pass_ticket, WX_Key

        platform_UserAgent = self.entries["platform_UserAgent"].get()
        platform_Cookie = self.entries["platform_Cookie"].get()
        platform_Token = self.entries["platform_Token"].get()
        platform_Fakeid = self.entries["platform_Fakeid"].get()
        WX_User_Agent = self.entries["WX_User_Agent"].get()
        WX_Cookie = self.entries["WX_Cookie"].get()
        WX_Appmsg_Token = self.entries["WX_Appmsg_Token"].get()
        WX_Pass_ticket = self.entries["WX_Pass_ticket"].get()
        WX_Key = self.entries["WX_Key"].get()

        # Redirect stdout to the Text widget
        sys.stdout = TextRedirector(self.output_area)

        main()

# Define the TextRedirector class to handle the stdout redirect
class TextRedirector(object):
    def __init__(self, widget):
        self.widget = widget

    def write(self, str):
        self.widget.insert(tk.END, str)
        self.widget.see(tk.END)

    def flush(self):
        pass


if __name__ == '__main__':
    GUI()
